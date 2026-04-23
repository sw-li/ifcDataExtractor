"""
exporter.py — Assembles extracted DataFrames and writes .xlsx output.

Sheet layout:
  Metadata    — one row per IFC file
  Hierarchy   — full spatial tree with element rows
  Psets       — all property values (long format)
  Quantities  — all quantity values

Export modes:
  per_ifc  — one .xlsx per input IFC, named {ifc_filename}.xlsx
  merged   — one .xlsx with a source_file column in every sheet, named merged.xlsx

Formatting:
  - Frozen header row
  - Auto column width
  - Alternating row shading via Excel Table (single XML operation, not per-cell)
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import Callable, Optional


# ---------------------------------------------------------------------------
# Shared style objects — created once at import time, reused for every cell
# ---------------------------------------------------------------------------
HEADER_FILL      = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT      = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_per_ifc(
    results: dict[str, dict[str, pd.DataFrame]],
    output_dir: str,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> list[str]:
    """
    Write one .xlsx file per IFC source.

    Parameters
    ----------
    results : dict
        {source_filename: {"Metadata": df, "Hierarchy": df, "Psets": df, "Quantities": df}}
    output_dir : str
        Folder where output files are written.
    progress_callback : callable | None
        Called with a status string after each file is written.

    Returns
    -------
    list[str]
        Paths of created files.
    """
    os.makedirs(output_dir, exist_ok=True)
    created = []

    for source_filename, dfs in results.items():
        base = os.path.splitext(os.path.basename(source_filename))[0]
        out_path = os.path.join(output_dir, f"{base}.xlsx")
        _write_workbook(dfs, out_path)
        created.append(out_path)
        if progress_callback:
            progress_callback(f"Written → {out_path}")

    return created


def export_merged(
    results: dict[str, dict[str, pd.DataFrame]],
    output_dir: str,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> str:
    """
    Write all IFC sources into a single merged .xlsx file.

    Parameters
    ----------
    results : dict
        {source_filename: {"Metadata": df, "Hierarchy": df, "Psets": df, "Quantities": df}}
    output_dir : str
    progress_callback : callable | None

    Returns
    -------
    str
        Path of the created file.
    """
    os.makedirs(output_dir, exist_ok=True)

    # Concatenate each sheet across all source files
    sheet_names = ["Metadata", "Hierarchy", "Psets", "Quantities"]
    merged: dict[str, pd.DataFrame] = {}

    for sheet in sheet_names:
        frames = [
            dfs[sheet]
            for dfs in results.values()
            if sheet in dfs and not dfs[sheet].empty
        ]
        merged[sheet] = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    out_path = os.path.join(output_dir, "merged.xlsx")
    _write_workbook(merged, out_path)

    if progress_callback:
        progress_callback(f"Written → {out_path}")

    return out_path


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_EXCEL_ROW_LIMIT = 1_048_575  # max data rows per sheet (Excel cap is 1 048 576 incl. header)


def _write_workbook(dfs: dict[str, pd.DataFrame], path: str) -> None:
    """Write a dict of DataFrames to an .xlsx workbook, one sheet each.

    If a DataFrame exceeds _EXCEL_ROW_LIMIT rows it is split across multiple
    sheets named "SheetName (1)", "SheetName (2)", etc. so no data is lost.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    sheet_order = ["Metadata", "Hierarchy", "Psets", "Quantities"]
    for sheet_name in sheet_order:
        df = dfs.get(sheet_name, pd.DataFrame())

        if len(df) <= _EXCEL_ROW_LIMIT:
            ws = wb.create_sheet(title=sheet_name)
            _write_sheet(ws, df)
        else:
            # Split into numbered chunks
            chunks = range(0, len(df), _EXCEL_ROW_LIMIT)
            total  = len(chunks)
            for part, start in enumerate(chunks, start=1):
                chunk = df.iloc[start : start + _EXCEL_ROW_LIMIT]
                title = f"{sheet_name} ({part})" if total > 1 else sheet_name
                ws = wb.create_sheet(title=title)
                _write_sheet(ws, chunk)

    wb.save(path)


def _write_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to an openpyxl worksheet with formatting.

    Alternating row shading is applied via an Excel Table (one XML entry)
    rather than a per-cell fill loop.  For a 50 000-row merged sheet this
    reduces formatting work from ~750 000 cell operations to essentially zero.
    """
    if df.empty:
        ws.append(["(no data)"])
        return

    headers  = list(df.columns)
    n_cols   = len(headers)

    # Track max column widths alongside the write loop — avoids a second
    # full scan of ws.columns after all data has been written.
    col_widths = [len(str(h)) + 4 for h in headers]

    # ── Header row ──────────────────────────────────────────────────────
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────────
    # Convert to plain Python lists once (faster than itertuples for large DFs).
    data_rows = df.to_numpy(dtype=object).tolist()

    for row_values in data_rows:
        ws.append(row_values)
        # Update column width estimates while we already have the row in memory
        for i, val in enumerate(row_values):
            if val is not None:
                w = len(str(val)) + 4
                if w > col_widths[i]:
                    col_widths[i] = w

    # ── Column widths ────────────────────────────────────────────────────
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)

    # ── Alternating row shading via Excel Table ───────────────────────────
    # A single Table entry in the workbook XML replaces the old O(rows × cols)
    # per-cell PatternFill loop.  Excel renders the stripe pattern natively.
    # Table names must be unique across all sheets in the workbook.
    last_col  = get_column_letter(n_cols)
    last_row  = len(df) + 1          # +1 for the header row
    table_ref = f"A1:{last_col}{last_row}"
    safe_name = "Tbl_" + "".join(c if c.isalnum() else "_" for c in ws.title)

    tbl = Table(displayName=safe_name, ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",   # blue header + light-blue / white stripes
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
